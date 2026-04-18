"""
conftest.py – Pytest fixtures dùng chung cho toàn bộ Selenium test suite
Quiz Trivia – https://datn-quizapp.web.app/
"""
import os
import pytest
import requests
import openpyxl
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# ── Load biến môi trường từ .env (tìm từ root project) ──────────────────────
_root = Path(__file__).resolve().parents[2]   # quiz-trivia-tooltest/
load_dotenv(_root / ".env")

# ── Firebase config ─────────────────────────────────────────────────────────
FIREBASE_API_KEY  = os.environ["FIREBASE_API_KEY"]
FIREBASE_PROJECT  = os.getenv("FIREBASE_PROJECT", "datn-quizapp")
FIRESTORE_BASE    = f"https://firestore.googleapis.com/v1/projects/{FIREBASE_PROJECT}/databases/(default)/documents"
SIGN_IN_URL       = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={FIREBASE_API_KEY}"

APP_URL = "https://datn-quizapp.web.app"

TEST_ACCOUNTS = {
    "user":    {"email": os.environ["USER_EMAIL"],    "password": os.environ["USER_PASSWORD"]},
    "creator": {"email": os.environ["CREATOR_EMAIL"], "password": os.environ["CREATOR_PASSWORD"]},
    "admin":   {"email": os.environ["ADMIN_EMAIL"],   "password": os.environ["ADMIN_PASSWORD"]},
}

# ── Selenium driver fixture ──────────────────────────────────────────────────
@pytest.fixture(scope="session")
def driver():
    """Headless Chrome driver – shared across the session."""
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-notifications")
    # webdriver_manager 4.x bug: install() may return THIRD_PARTY_NOTICES instead
    # of chromedriver.exe. Resolve the real executable from the same directory.
    raw_path = ChromeDriverManager().install()
    driver_path = Path(raw_path)
    if driver_path.suffix != ".exe" and not driver_path.name.startswith("chromedriver"):
        candidate = driver_path.parent / "chromedriver.exe"
        if candidate.exists():
            raw_path = str(candidate)
    service = Service(raw_path)
    drv = webdriver.Chrome(service=service, options=opts)
    drv.implicitly_wait(10)
    yield drv
    drv.quit()


# ── Firebase helpers ─────────────────────────────────────────────────────────
def get_id_token(role: str = "user") -> dict:
    """Sign in and return {idToken, localId}."""
    acc = TEST_ACCOUNTS[role]
    resp = requests.post(SIGN_IN_URL, json={
        "email": acc["email"],
        "password": acc["password"],
        "returnSecureToken": True,
    })
    resp.raise_for_status()
    data = resp.json()
    return {"idToken": data["idToken"], "localId": data["localId"]}


def firestore_get(collection: str, doc_id: str, id_token: str) -> dict | None:
    """GET a Firestore document. Returns None if not found."""
    url = f"{FIRESTORE_BASE}/{collection}/{doc_id}"
    headers = {"Authorization": f"Bearer {id_token}"}
    resp = requests.get(url, headers=headers)
    if resp.status_code == 404:
        return None
    resp.raise_for_status()
    return resp.json()


def firestore_delete(collection: str, doc_id: str, id_token: str):
    """DELETE a Firestore document (rollback helper)."""
    url = f"{FIRESTORE_BASE}/{collection}/{doc_id}"
    headers = {"Authorization": f"Bearer {id_token}"}
    resp = requests.delete(url, headers=headers)
    if resp.status_code not in (200, 204, 404):
        print(f"  [WARN] Rollback delete failed: {collection}/{doc_id} → {resp.status_code}")


# ── Excel result writer ──────────────────────────────────────────────────────
class ExcelResultWriter:
    """Ghi kết quả Pass/Fail + Actual ngược vào file Excel (theo hướng dẫn cô)."""

    def __init__(self, xlsx_path: str, sheet_name: str,
                 tc_col=1, status_col=12,
                 actual_ui_col=10, actual_db_col=11,
                 data_start_row=8):
        self.path          = xlsx_path
        self.sheet_name    = sheet_name
        self.tc_col        = tc_col
        self.status_col    = status_col
        self.actual_ui_col = actual_ui_col
        self.actual_db_col = actual_db_col
        self.start_row     = data_start_row
        self.wb = openpyxl.load_workbook(xlsx_path)
        self.ws = self.wb[sheet_name]

    def _find_row(self, tc_id: str) -> int | None:
        for row in self.ws.iter_rows(min_row=self.start_row):
            if str(row[self.tc_col - 1].value).strip() == tc_id.strip():
                return row[0].row
        return None

    def write(self, tc_id: str, status: str,
              actual_ui: str = "", actual_db: str = ""):
        row = self._find_row(tc_id)
        if row is None:
            print(f"  [WARN] TC not found in Excel: {tc_id}")
            return
        self.ws.cell(row, self.status_col, status)
        if actual_ui:
            self.ws.cell(row, self.actual_ui_col, actual_ui)
        if actual_db:
            self.ws.cell(row, self.actual_db_col, actual_db)
        # timestamp in Notes column (last col)
        note_col = self.ws.max_column
        existing = self.ws.cell(row, note_col).value or ""
        self.ws.cell(row, note_col,
                     f"{existing} | Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        self.wb.save(self.path)


@pytest.fixture(scope="session")
def excel_login():
    return ExcelResultWriter(
        xlsx_path="../../data/TC_Login.xlsx",
        sheet_name="TC_Login",
        status_col=12, actual_ui_col=10, actual_db_col=11, data_start_row=8,
    )


@pytest.fixture(scope="session")
def excel_quiz_core():
    return ExcelResultWriter(
        xlsx_path="../../data/TC_QuizCoreFlow.xlsx",
        sheet_name="TC_QuizCoreFlow",
        status_col=11, actual_ui_col=9, actual_db_col=10, data_start_row=8,
    )


@pytest.fixture(scope="session")
def excel_quiz_mgmt():
    return ExcelResultWriter(
        xlsx_path="../../data/TC_QuizManagement.xlsx",
        sheet_name="TC_QuizCRUD",
        status_col=11, actual_ui_col=9, actual_db_col=10, data_start_row=6,
    )


@pytest.fixture(scope="session")
def excel_ai():
    return ExcelResultWriter(
        xlsx_path="../../data/TC_AIFeatures.xlsx",
        sheet_name="TC_AIGenerator",
        status_col=11, actual_ui_col=8, actual_db_col=9, data_start_row=6,
    )
